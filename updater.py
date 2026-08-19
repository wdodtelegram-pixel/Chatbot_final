"""
updater.py  (OPTIONAL)
======================
PURPOSE
    Lets authorised FM Officers update job records directly from Telegram
    without opening Excel at all, using commands like:

        /update ACMV-001 status Completed
        /update ACMV-001 est_completion Week 2 February 2026
        /update ACMV-001 next_job Week 2 May 2026

    Changes are written immediately to the local Excel file. If OneDrive
    desktop sync is running, the change propagates back to Excel Online
    automatically -- so the spreadsheet stays the source of truth in both
    directions.

WIRING INTO bot.py
    1. Add your Telegram user ID to AUTHORIZED_USER_IDS below.
       (Send a message to @userinfobot to find yours.)
    2. In bot.py, after the other imports, add:
           from updater import update_command
    3. In main(), before the catch-all handler, add:
           application.add_handler(CommandHandler("update", update_command))
    4. In post_init(), add to the commands list:
           BotCommand("update", "Update a job field (FM Officers only)")
"""

from __future__ import annotations
import logging

import openpyxl
from telegram import Update
from telegram.constants import ParseMode
from telegram.ext import ContextTypes

import config
from date_utils import to_week_format

logger = logging.getLogger(__name__)

# -----------------------------------------------------------------------
# Add the Telegram user IDs of people allowed to write data.
# Everyone else gets a silent refusal -- no error, because making it
# obvious that an /update command exists would invite guessing.
# Find your ID by messaging @userinfobot on Telegram.
# -----------------------------------------------------------------------
AUTHORIZED_USER_IDS: set[int] = {
    # 123456789,   # <- replace with your real Telegram user ID
}

# Which field names the command accepts -> which Excel column they map to.
FIELD_MAP = {
    "status":         config.COL_STATUS,
    "est_completion": config.COL_EST_COMPLETION,
    "completion":     config.COL_EST_COMPLETION,
    "next_job":       config.COL_NEXT_JOB,
    "location":       config.COL_LOCATION,
    "type":           config.COL_TYPE,
}

# Allowed values for fields that have a fixed set, checked before writing.
ALLOWED_VALUES = {
    config.COL_STATUS: {"Yet to Start", "In Progress", "Completed"},
    config.COL_TYPE:   {"Preventive", "Corrective"},
}


def _parse_command(args: list[str]) -> tuple[str, str, str] | None:
    """
    PURPOSE: parse the three parts of the command from its arguments.

    /update ACMV-001 status Completed
    args = ["ACMV-001", "status", "Completed"]
    returns ("ACMV-001", "status", "Completed")

    /update ACMV-001 est_completion Week 2 February 2026
    args = ["ACMV-001", "est_completion", "Week", "2", "February", "2026"]
    returns ("ACMV-001", "est_completion", "Week 2 February 2026")
    """
    if len(args) < 3:
        return None
    job_id = args[0].upper()
    field  = args[1].lower()
    value  = " ".join(args[2:]).strip()
    return job_id, field, value


def _write_to_excel(job_id: str, col_name: str, new_value: str) -> str | None:
    """
    PURPOSE: open the workbook, find the row, update the cell, save.

    Returns an error string on failure, or None on success.

    openpyxl is used directly (not pandas) because pandas re-reads the whole
    file into memory and writing back requires a full round-trip that can lose
    formatting and validation. openpyxl edits the file in place.
    """
    path = config.EXCEL_PATH
    if not path.exists():
        return f"Workbook not found at {path}."

    try:
        wb = openpyxl.load_workbook(path)
    except Exception as exc:
        return f"Could not open workbook: {exc}"

    if config.SHEET_NAME not in wb.sheetnames:
        return f"Sheet '{config.SHEET_NAME}' not found."

    ws = wb[config.SHEET_NAME]

    # Find the column index for the header we want.
    header_row = [cell.value for cell in ws[1]]
    if col_name not in header_row:
        return f"Column '{col_name}' not found in the sheet."
    col_index = header_row.index(col_name) + 1  # openpyxl is 1-based

    # Find the Job ID column.
    if config.COL_JOB_ID not in header_row:
        return "Job ID column not found."
    id_col_index = header_row.index(config.COL_JOB_ID) + 1

    # Search for the row matching job_id.
    target_row = None
    for row in ws.iter_rows(min_row=2):
        cell_val = row[id_col_index - 1].value
        if str(cell_val).strip().upper() == job_id:
            target_row = row
            break

    if target_row is None:
        return f"Job ID '{job_id}' not found in the sheet."

    # Write and save.
    ws.cell(row=target_row[0].row, column=col_index, value=new_value)
    try:
        wb.save(path)
    except PermissionError:
        return (
            "The file is locked (probably open in Excel). "
            "Close it and try again."
        )
    except Exception as exc:
        return f"Could not save workbook: {exc}"

    return None  # success


async def update_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """
    PURPOSE: the /update command handler. Validates auth, parses args, writes.

    Usage:
        /update <Job-ID> <field> <value>

    Examples:
        /update ACMV-001 status Completed
        /update LFT-002 est_completion Week 3 February 2026
        /update ELE-003 location Level 6 Server Room
    """
    # --- auth ---
    user = update.effective_user
    if not user or user.id not in AUTHORIZED_USER_IDS:
        # No reply -- don't advertise that /update exists.
        return

    args = context.args or []
    parsed = _parse_command(args)

    if parsed is None:
        await update.message.reply_text(
            "Usage: <code>/update &lt;Job-ID&gt; &lt;field&gt; &lt;value&gt;</code>\n\n"
            "Fields: status, est_completion, next_job, location, type\n\n"
            "Examples:\n"
            "<code>/update ACMV-001 status Completed</code>\n"
            "<code>/update ACMV-001 est_completion Week 3 February 2026</code>",
            parse_mode=ParseMode.HTML,
        )
        return

    job_id, field_key, value = parsed

    # --- validate field name ---
    if field_key not in FIELD_MAP:
        await update.message.reply_text(
            f"Unknown field <code>{field_key}</code>.\n"
            f"Allowed: {', '.join(sorted(FIELD_MAP))}",
            parse_mode=ParseMode.HTML,
        )
        return

    col_name = FIELD_MAP[field_key]

    # --- normalise value against canonical maps ---
    if col_name == config.COL_STATUS:
        value = config.STATUS_CANONICAL.get(value.lower(), value)
    elif col_name == config.COL_TYPE:
        value = config.TYPE_CANONICAL.get(value.lower(), value)

    # --- validate against allowed sets ---
    if col_name in ALLOWED_VALUES and value not in ALLOWED_VALUES[col_name]:
        allowed = ", ".join(sorted(ALLOWED_VALUES[col_name]))
        await update.message.reply_text(
            f"<code>{value}</code> is not valid for <b>{field_key}</b>.\n"
            f"Allowed values: {allowed}",
            parse_mode=ParseMode.HTML,
        )
        return

    # --- write ---
    error = _write_to_excel(job_id, col_name, value)
    if error:
        await update.message.reply_text(f"⚠️ {error}", parse_mode=ParseMode.HTML)
        return

    # Force the loader cache to pick up the change we just made.
    import excel_loader
    excel_loader.load_jobs(force=True)

    display_value = to_week_format(value) if "date" in field_key or "completion" in field_key else value
    await update.message.reply_text(
        f"✅ <b>{job_id}</b> updated.\n"
        f"   {field_key} → <code>{display_value}</code>",
        parse_mode=ParseMode.HTML,
    )
    logger.info("User %d updated %s.%s = %r", user.id, job_id, col_name, value)
